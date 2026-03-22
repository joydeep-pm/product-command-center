from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
DOCS_DIR = ROOT / "docs"

SOURCE_FILES = {
    "strategy_deck": Path("/Users/joy/Downloads/M2P Finflux CLS FY27 Strategy Deck v1.0.pdf"),
    "strategy_report": Path("/Users/joy/Downloads/Core Lending Strategy FY27 Report v1.0.pdf"),
    "roadmap_pdf": Path("/Users/joy/Downloads/Lending Roadmap (1).pdf"),
    "roadmap_xlsx": Path("/Users/joy/Downloads/Core Lending FY27 Roadmap Master v1.0.xlsx"),
}


def clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        return " ".join(value.split())
    return value


def row_dict(headers: list[Any], row: list[Any]) -> dict[str, Any]:
    item: dict[str, Any] = {}
    for key, value in zip(headers, row):
        if key is None:
            continue
        key_clean = clean(key)
        if not key_clean:
            continue
        item[str(key_clean)] = clean(value)
    return item


def load_sheet_records(ws, header_row: int, start_row: int) -> list[dict[str, Any]]:
    headers = [cell for cell in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
    records = []
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        item = row_dict(headers, list(row))
        if item:
            records.append(item)
    return records


def extract_workbook(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True)

    roadmap_master = load_sheet_records(wb["1_ROADMAP_MASTER"], 3, 4)
    q_epics = load_sheet_records(wb["2_Q_BY_Q_EPICS"], 3, 5)
    loss_intel = load_sheet_records(wb["3_LOSS_INTELLIGENCE"], 3, 4)
    parity = load_sheet_records(wb["4_COMPETITIVE_PARITY"], 3, 4)
    resource_plan = load_sheet_records(wb["5_RESOURCE_PLAN"], 3, 4)
    topic_map = load_sheet_records(wb["6_TOPIC_MAP"], 4, 5)

    csuite_pairs: dict[str, list[str]] = defaultdict(list)
    ws = wb["7_CSUITE_SNAPSHOT"]
    for row in ws.iter_rows(min_row=3, values_only=True):
        values = [clean(v) for v in row]
        for i in range(0, len(values), 2):
            label = values[i]
            metric = values[i + 1] if i + 1 < len(values) else None
            if label:
                if metric:
                    csuite_pairs[str(label)].append(str(metric))

    quarter_summary: dict[str, dict[str, Any]] = defaultdict(lambda: {"epics": 0, "pd_total": 0, "priorities": defaultdict(int)})
    for item in roadmap_master:
        quarter = item.get("Quarter")
        if not quarter or quarter == "Ongoing":
            continue
        q = quarter_summary[str(quarter)]
        q["epics"] += 1
        pd_val = item.get("PD (Revised/Added)") or item.get("PD (Original AOP)")
        if isinstance(pd_val, (int, float)):
            q["pd_total"] += int(pd_val)
        priority = item.get("Dev Priority")
        if priority:
            q["priorities"][str(priority)] += 1

    for q, info in quarter_summary.items():
        info["priorities"] = dict(info["priorities"])

    workbook_loss_total = round(
        sum(item.get("Loss (₹ Cr)") or 0 for item in loss_intel if isinstance(item.get("Loss (₹ Cr)"), (int, float))),
        1,
    )
    workbook_quarter_pd_total = sum(info["pd_total"] for info in quarter_summary.values())
    workbook_quarter_epic_total = sum(info["epics"] for info in quarter_summary.values())

    executive_baseline = {
        "declared_fy26_losses_inr_cr": 100.5,
        "workbook_loss_total_inr_cr": workbook_loss_total,
        "demo_loss_rate_pct": 78.5,
        "recoverable_pipeline_inr_cr_range": "49-55",
        "net_new_pipeline_inr_cr_range": "25-40",
        "ai_acv_uplift_pct_range": "15-25",
        "roadmap_versions": {
            "strategy_deck_pd": 5551,
            "strategy_deck_epics": 42,
            "strategy_report_pd": 5385,
            "strategy_report_epics": 40,
            "roadmap_pdf_pd": 4368,
            "workbook_quarter_sum_pd": workbook_quarter_pd_total,
            "workbook_quarter_sum_epics": workbook_quarter_epic_total,
        },
        "data_quality_notes": [
            "The workbook row-level loss sum does not equal the 100.5 Cr headline from the deck/report.",
            "Roadmap totals vary across deck, report, roadmap PDF, and workbook line-item sums.",
            "The command center should surface these as versioned planning numbers, not silently collapse them into one truth.",
        ],
    }

    return {
        "source_summary": {
            "workbook": str(path),
            "sheet_names": wb.sheetnames,
        },
        "executive_baseline": executive_baseline,
        "csuite_snapshot": dict(csuite_pairs),
        "roadmap_master": roadmap_master,
        "quarterly_epics": q_epics,
        "loss_intelligence": loss_intel,
        "competitive_parity": parity,
        "resource_plan": resource_plan,
        "topic_map": topic_map,
        "quarter_summary": dict(quarter_summary),
    }


PDF_HIGHLIGHTS = {
    "strategy_deck": [
        "Executive Summary",
        "Financial Impact & ROI Framework",
        "Strategic Bet: Islamic Finance & GCC, ASEAN Market Entry",
        "Domain Specialist Hiring",
    ],
    "strategy_report": [
        "Executive Summary",
        "Market & Competitive Landscape",
        "Financial Impact & ROI Framework",
        "Recommendation",
    ],
    "roadmap_pdf": [
        "Q1",
        "Q2",
        "Q3",
        "Q4",
        "Grand Total",
    ],
}


def extract_pdf_notes(path: Path, key: str) -> dict[str, Any]:
    reader = PdfReader(str(path))
    full_text = []
    for page in reader.pages:
        full_text.append(page.extract_text() or "")
    text = "\n".join(full_text)
    snippets = []
    for needle in PDF_HIGHLIGHTS.get(key, []):
        idx = text.find(needle)
        if idx >= 0:
            snippets.append(text[idx: idx + 900].replace("\n", " "))
    return {
        "path": str(path),
        "pages": len(reader.pages),
        "highlights": snippets,
    }


def build_strategic_summary(data: dict[str, Any]) -> dict[str, Any]:
    executive = data["executive_baseline"]
    q_summary = data["quarter_summary"]
    q1_epics = q_summary.get("Q1", {}).get("epics", 0)
    q2_epics = q_summary.get("Q2", {}).get("epics", 0)
    q3_epics = q_summary.get("Q3", {}).get("epics", 0)
    loss_total = executive["declared_fy26_losses_inr_cr"]
    return {
        "imperatives": [
            {
                "title": "Close the Revenue Bleed",
                "color": "var(--red)",
                "body": f"₹{loss_total}+ Cr lost in FY26. The source docs repeatedly concentrate the fix on Gold (20% parity), LAMF/LAS (53%), BNPL/Credit Line (71%), and Vehicle Finance. Q1 carries {q1_epics} workbook EPICs to recover the highest-confidence lost revenue first.",
                "source_refs": ["strategy_deck", "strategy_report", "competitive_parity", "loss_intelligence", "quarter_summary.Q1"],
            },
            {
                "title": "Pass the Compliance Firewall",
                "color": "var(--amber)",
                "body": f"DPDPA, ECL/EIR, Audit Module 2.0, and NPA automation show up as veto gates across bank and large NBFC deals. Q2 has {q2_epics} workbook EPICs, with compliance-heavy scope positioned to remove those gates before larger enterprise pushes.",
                "source_refs": ["strategy_deck", "strategy_report", "roadmap_master", "quarter_summary.Q2"],
            },
            {
                "title": "Enter Two New High-Value Markets",
                "color": "var(--blue)",
                "body": f"Islamic Finance for GCC/ASEAN and Capital Markets/B2B are the clearest net-new FY27 bets in the source docs. Q3 concentrates {q3_epics} workbook EPICs into those wedges instead of treating them as side quests.",
                "source_refs": ["strategy_deck", "strategy_report", "roadmap_master", "quarter_summary.Q3"],
            },
            {
                "title": "Accelerate Delivery with AI Tooling",
                "color": "var(--green)",
                "body": "The source narrative treats AI twice: internal engineering leverage and external product monetization. The current executive baseline assumes 1,100-1,400 PD of engineering recovery plus 15-25% AI-linked ACV uplift when packaged correctly.",
                "source_refs": ["strategy_deck", "strategy_report", "executive_baseline"],
            },
            {
                "title": "Build the Loss-Prevention Framework",
                "color": "var(--purple)",
                "body": "The report does not frame FY26 losses as feature-only. Trust deficit, weak references, implementation friction, and poor qualification all recur. Hypercare, QBRs, tighter qualification, and win-loss loops are therefore operating requirements, not GTM decoration.",
                "source_refs": ["strategy_report", "loss_intelligence", "topic_map"],
            },
        ],
        "moats": [
            {
                "title": "Islamic Finance / GCC",
                "color": "var(--amber)",
                "icon": "🏅",
                "body": "The strategy deck calls this the only category where Finflux can plausibly create first-mover advantage in FY27: cloud-native Islamic LMS capability for GCC/ASEAN with Shariah governance, not just conventional lending relabeled.",
                "source_refs": ["strategy_deck", "strategy_report", "competitive_parity"],
            },
            {
                "title": "AI-Native LMS Intelligence",
                "color": "var(--blue)",
                "icon": "🤖",
                "body": "The moat is not generic AI. It comes from portfolio, repayment, collections, and covenant signals already living inside LMS workflows. The source docs connect that to both internal leverage and premium ACV expansion.",
                "source_refs": ["strategy_deck", "strategy_report", "executive_baseline"],
            },
            {
                "title": "Co-lending Orchestration Hub",
                "color": "var(--green)",
                "icon": "⚡",
                "body": "Co-lending is treated as infrastructure, not one feature. Middleware, triple-book accounting, partner onboarding, and repeatable integrations create a distribution and delivery moat if Q1 execution is credible enough to become referenceable.",
                "source_refs": ["strategy_deck", "strategy_report", "roadmap_master", "topic_map"],
            },
        ],
        "friction_zones": [
            {
                "label": "01 Product",
                "title": "Product Maturity & Gaps",
                "color": "var(--red)",
                "body": "The sources repeatedly position Finflux as evaluation-worthy but not enterprise-complete. Gold, LAMF, BNPL, compliance depth, and collections maturity are the most repeated examples.",
                "source_refs": ["strategy_report", "competitive_parity", "loss_intelligence"],
            },
            {
                "label": "02 Engineering",
                "title": "Velocity & Innovation Dilemma",
                "color": "var(--amber)",
                "body": "Escalations, stabilization work, and implementation drag absorb roadmap energy. The FY27 plan responds with bug-bash cycles, AI leverage, and sharper sequencing rather than pretending velocity is already solved.",
                "source_refs": ["strategy_report", "roadmap_master", "resource_plan"],
            },
            {
                "label": "03 GTM",
                "title": "Sales Execution Misalignment",
                "color": "var(--blue)",
                "body": "The strategy report is explicit that qualification, demo fidelity, and wishlist selling are contributing to losses. That means product truth and sales discipline need to tighten together.",
                "source_refs": ["strategy_report", "loss_intelligence", "executive_baseline"],
            },
            {
                "label": "04 Ecosystem",
                "title": "Ecosystem Stagnation",
                "color": "var(--purple)",
                "body": "Integration and implementation repeatability are still weak points. The plan treats SI readiness, co-lending infrastructure, and reusable integration assets as multipliers on product investment.",
                "source_refs": ["strategy_deck", "strategy_report", "topic_map"],
            },
            {
                "label": "05 LOS/BRE",
                "title": "LOS Foundation (WIP)",
                "color": "var(--cyan)",
                "body": "The source narrative keeps returning to orchestration quality: LOS, BRE, config, and integration spine are not optional plumbing. Weakness there degrades every downstream product category and implementation.",
                "source_refs": ["strategy_report", "roadmap_master", "topic_map"],
            },
        ],
        "key_insight": {
            "label": "Key Insight",
            "title": "Completion Problem, Not Architecture",
            "color": "var(--green)",
            "body": "The strategy report's 'Almost Comply' framing is the clearest synthesis: Finflux gets evaluated because the platform is credible, then loses the final 10-20% on completeness, compliance depth, and trust. FY27 is therefore a completion program, not a platform rewrite.",
            "source_refs": ["strategy_report", "competitive_parity", "loss_intelligence"],
        },
    }


def build_source_brief(data: dict[str, Any], pdf_notes: dict[str, Any]) -> str:
    q_summary = data["quarter_summary"]
    executive = data["executive_baseline"]
    top_losses = data["loss_intelligence"][:8]
    top_gaps = [item for item in data["competitive_parity"] if item.get("Gap Severity") in {"Critical", "CRITICAL", "High"}][:8]
    lines = [
        "# Source Brief",
        "",
        "This brief condenses the source documents into the command-center content model.",
        "",
        "## Executive spine",
        "",
        f"- FY26 losses headline: Rs {executive['declared_fy26_losses_inr_cr']} Cr.",
        f"- Workbook row-level loss sum: Rs {executive['workbook_loss_total_inr_cr']} Cr.",
        f"- Demo loss rate: {executive['demo_loss_rate_pct']}%.",
        "- Primary parity gaps called out repeatedly: Gold, LAS/LAMF, BNPL, compliance, delivery trust.",
        "- Strategic sequence: Q1 parity closure, Q2 compliance readiness, Q3 new markets, Q4 ecosystem monetization.",
        "- Treat roadmap and loss totals as versioned planning numbers when the source documents disagree.",
        "",
        "## Quarter summary",
        "",
    ]
    for quarter in sorted(q_summary):
        info = q_summary[quarter]
        lines.append(f"- {quarter}: {info['epics']} epics, {info['pd_total']} PD, priority mix {info['priorities']}")
    lines.extend([
        "",
        "## Top loss intelligence rows",
        "",
    ])
    for item in top_losses:
        lines.append(
            f"- {item.get('Prospect')}: loss Rs {item.get('Loss (₹ Cr)')} Cr, product {item.get('Product')}, reason {item.get('Primary Reason')}, mapped fix {item.get('Mapped EPIC(s)')}, quarter {item.get('Quarter Fix')}"
        )
    lines.extend([
        "",
        "## Critical parity gaps",
        "",
    ])
    for item in top_gaps:
        lines.append(
            f"- {item.get('Product')} / {item.get('Feature Group')}: Finflux status {item.get('Finflux Status')}, severity {item.get('Gap Severity')}, mapped epic {item.get('Mapped EPIC')}, target quarter {item.get('Q Fix')}"
        )
    lines.extend([
        "",
        "## PDF framing notes",
        "",
    ])
    for key, note in pdf_notes.items():
        lines.append(f"### {key}")
        lines.append("")
        lines.append(f"- Pages: {note['pages']}")
        for snippet in note["highlights"][:4]:
            lines.append(f"- {snippet[:300]}...")
        lines.append("")
    lines.extend([
        "## Required dashboard sections",
        "",
        "- Executive snapshot",
        "- Strategic imperatives",
        "- Quarter-by-quarter roadmap",
        "- Epic register",
        "- Loss intelligence",
        "- Competitive parity",
        "- Compliance firewall",
        "- Market expansion bets",
        "- Resource plan",
        "- Risks and dependencies",
        "- Source index",
        "",
    ])
    return "\n".join(lines)


def main() -> None:
    DATA_DIR.mkdir(exist_ok=True)
    DOCS_DIR.mkdir(exist_ok=True)

    workbook_data = extract_workbook(SOURCE_FILES["roadmap_xlsx"])
    workbook_data["strategic_summary"] = build_strategic_summary(workbook_data)
    pdf_notes = {
        key: extract_pdf_notes(path, key)
        for key, path in SOURCE_FILES.items()
        if path.suffix.lower() == ".pdf"
    }
    payload = {
        "meta": {
            "fiscal_year": "FY 2026-27",
            "product": "M2P Finflux Core Lending Platform",
            "sources": {k: str(v) for k, v in SOURCE_FILES.items()},
        },
        "workbook": workbook_data,
        "pdf_notes": pdf_notes,
    }
    (DATA_DIR / "command_center_data.json").write_text(json.dumps(payload, indent=2, ensure_ascii=True))
    (DOCS_DIR / "source_brief.md").write_text(build_source_brief(workbook_data, pdf_notes))
    print(f"Wrote {(DATA_DIR / 'command_center_data.json')}")
    print(f"Wrote {(DOCS_DIR / 'source_brief.md')}")


if __name__ == "__main__":
    main()
